"""
cims_gui.py — Consumables Inventory Management System for IsoWorks.

Three-tab widget:
  • Inventory   — item master + lot sub-table (CRUD)
  • Usage Log   — read-only movement history with filters and CSV export
  • Reports     — stock valuation and usage-per-run export (Excel / PDF)
"""
from __future__ import annotations

import csv
import getpass
import logging
import os
from datetime import datetime, date, timedelta
from typing import Optional

from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QGroupBox, QSplitter, QTabWidget, QDateEdit,
    QDialog, QDialogButtonBox, QDoubleSpinBox, QTextEdit,
    QFileDialog, QMessageBox, QSizePolicy, QFrame,
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtGui import QTextDocument
from sqlalchemy import text

from db_core import db_manager
from gui_utils import show_message
from help_browser import make_help_button
from shared_utils import get_global_value

log = logging.getLogger(__name__)

# ── palette ──────────────────────────────────────────────────────────────────
_RED    = QColor("#FFCDD2")
_AMBER  = QColor("#FFF9C4")
_GREEN  = QColor("#E8F5E9")
_GREY   = QColor("#ECEFF1")
_BLUE   = QColor("#E3F2FD")

_CATEGORIES   = ["Chemical", "LabSupply", "Gas", "Instrument"]
_MOVE_TYPES   = ["RECEIVED", "USED", "ADJUSTED", "SCRAPPED"]
_MODULES      = ["SIAM", "TRIMS", "NGAM"]
_MOVE_COLOURS = {
    "RECEIVED": _GREEN,
    "USED":     QColor("#FFFFFF"),
    "ADJUSTED": _BLUE,
    "SCRAPPED": _RED,
}

_HDR_SS = (
    "QHeaderView::section {"
    "  background:#37474F; color:white; font-weight:bold;"
    "  padding:4px 6px; border:none;"
    "}"
)
_BTN_SS = (
    "QPushButton{background:#546E7A;color:white;font-weight:bold;"
    "border:none;padding:4px 10px;border-radius:3px;}"
    "QPushButton:hover{background:#37474F;}"
    "QPushButton:disabled{background:#B0BEC5;color:#78909C;}"
)
_BTN_ADD_SS = (
    "QPushButton{background:#2E7D32;color:white;font-weight:bold;"
    "border:none;padding:4px 10px;border-radius:3px;}"
    "QPushButton:hover{background:#1B5E20;}"
    "QPushButton:disabled{background:#B0BEC5;color:#78909C;}"
)
_BTN_DEL_SS = (
    "QPushButton{background:#C62828;color:white;font-weight:bold;"
    "border:none;padding:4px 10px;border-radius:3px;}"
    "QPushButton:hover{background:#B71C1C;}"
    "QPushButton:disabled{background:#B0BEC5;color:#78909C;}"
)


def _tbl(headers: list[str], parent=None) -> QTableWidget:
    """Return a read-only, selection-by-row QTableWidget."""
    t = QTableWidget(0, len(headers), parent)
    t.setHorizontalHeaderLabels(headers)
    t.horizontalHeader().setStyleSheet(_HDR_SS)
    t.horizontalHeader().setStretchLastSection(True)
    t.verticalHeader().setVisible(False)
    t.setEditTriggers(QAbstractItemView.NoEditTriggers)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setAlternatingRowColors(True)
    t.setSortingEnabled(True)
    return t


def _cell(text: str, align=Qt.AlignLeft) -> QTableWidgetItem:
    it = QTableWidgetItem(str(text) if text is not None else "")
    it.setTextAlignment(align | Qt.AlignVCenter)
    return it


# ── Lot dialog ───────────────────────────────────────────────────────────────

class _LotDialog(QDialog):
    """Add a new lot / shipment for an item."""

    def __init__(self, item_id: int, item_name: str, suppliers: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Add Lot — {item_name}")
        self.setMinimumWidth(420)
        self._item_id = item_id
        self._suppliers = suppliers  # [(supplier_id, name), …]

        f = QFormLayout(self)
        f.setContentsMargins(12, 12, 12, 8)
        f.setSpacing(8)

        self.txtLot     = QLineEdit()
        self.cmbSupp    = QComboBox()
        self.cmbSupp.addItem("— none —", None)
        for sid, sname in suppliers:
            self.cmbSupp.addItem(sname, sid)

        self.dtReceived = QDateEdit(QDate.currentDate())
        self.dtReceived.setCalendarPopup(True)
        self.dtExpiry   = QDateEdit()
        self.dtExpiry.setCalendarPopup(True)
        self.dtExpiry.setSpecialValueText("(no expiry)")
        self.dtExpiry.setDate(QDate(2099, 12, 31))

        self.spnQty     = QDoubleSpinBox(); self.spnQty.setRange(0, 1e9); self.spnQty.setDecimals(3)
        self.spnCost    = QDoubleSpinBox(); self.spnCost.setRange(0, 1e9); self.spnCost.setDecimals(4)
        self.txtInvoice = QLineEdit()
        self.txtRemarks = QLineEdit()

        f.addRow("Lot / batch #:", self.txtLot)
        f.addRow("Supplier:",      self.cmbSupp)
        f.addRow("Date received:", self.dtReceived)
        f.addRow("Expiry date:",   self.dtExpiry)
        f.addRow("Quantity:",      self.spnQty)
        f.addRow("Unit cost:",     self.spnCost)
        f.addRow("Invoice ref:",   self.txtInvoice)
        f.addRow("Remarks:",       self.txtRemarks)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self._on_ok)
        bb.rejected.connect(self.reject)
        f.addRow(bb)

    def _on_ok(self):
        if self.spnQty.value() <= 0:
            show_message(self, "Validation", "Quantity must be greater than zero.")
            return
        self.accept()

    def values(self) -> dict:
        exp = self.dtExpiry.date()
        expiry = date(exp.year(), exp.month(), exp.day()) if exp.year() < 2099 else None
        recv   = self.dtReceived.date()
        return {
            "item_id":      self._item_id,
            "supplier_id":  self.cmbSupp.currentData(),
            "lot_number":   self.txtLot.text().strip() or None,
            "date_received": date(recv.year(), recv.month(), recv.day()),
            "expiry_date":  expiry,
            "qty_received": self.spnQty.value(),
            "qty_remaining": self.spnQty.value(),
            "unit_cost":    self.spnCost.value() or None,
            "invoice_ref":  self.txtInvoice.text().strip() or None,
            "remarks":      self.txtRemarks.text().strip() or None,
        }


# ── Usage dialog ─────────────────────────────────────────────────────────────

class _UsageDialog(QDialog):
    """Record a stock movement (USED / ADJUSTED / SCRAPPED) against a lot."""

    def __init__(self, lot_id: int, lot_label: str, max_qty: float, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Record Usage — {lot_label}")
        self.setMinimumWidth(380)
        self._lot_id = lot_id

        f = QFormLayout(self)
        f.setContentsMargins(12, 12, 12, 8)
        f.setSpacing(8)

        self.cmbType    = QComboBox()
        for t in ["USED", "ADJUSTED", "SCRAPPED"]:
            self.cmbType.addItem(t, t)

        self.spnQty     = QDoubleSpinBox()
        self.spnQty.setRange(0.001, max_qty if max_qty > 0 else 1e9)
        self.spnQty.setDecimals(3)
        self.spnQty.setValue(min(1.0, max_qty))

        self.cmbModule  = QComboBox()
        self.cmbModule.addItem("— none —", None)
        for m in _MODULES:
            self.cmbModule.addItem(m, m)

        self.spnRunID   = QDoubleSpinBox()
        self.spnRunID.setRange(0, 2_000_000)
        self.spnRunID.setDecimals(0)
        self.spnRunID.setSpecialValueText("—")

        self.txtPurpose = QLineEdit()

        f.addRow("Movement type:", self.cmbType)
        f.addRow("Quantity:",      self.spnQty)
        f.addRow("Run module:",    self.cmbModule)
        f.addRow("Run ID:",        self.spnRunID)
        f.addRow("Purpose:",       self.txtPurpose)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        f.addRow(bb)

    def values(self) -> dict:
        run_id = int(self.spnRunID.value()) if self.spnRunID.value() > 0 else None
        return {
            "lot_id":        self._lot_id,
            "movement_type": self.cmbType.currentData(),
            "quantity":      self.spnQty.value(),
            "used_by":       getpass.getuser(),
            "run_module":    self.cmbModule.currentData(),
            "run_id":        run_id,
            "purpose":       self.txtPurpose.text().strip() or None,
        }


# ── Inventory panel ───────────────────────────────────────────────────────────

class CIMSInventoryPanel(QWidget):
    """Left: item list with colour-coded stock status. Right: detail + lot sub-table."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._current_item_id: Optional[int] = None
        self._is_new = False
        self._suppliers: list = []  # [(id, name), …]

        splitter = QSplitter(Qt.Horizontal, self)

        # ── Left: list ──────────────────────────────────────────────────────
        left = QWidget()
        lv   = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 4, 0)

        # filter / search bar
        fb = QHBoxLayout()
        fb.addWidget(QLabel("Category:"))
        self._cmbCat = QComboBox(); self._cmbCat.addItem("All", None)
        for c in _CATEGORIES: self._cmbCat.addItem(c, c)
        fb.addWidget(self._cmbCat)
        fb.addWidget(QLabel("Search:"))
        self._txtSearch = QLineEdit(); self._txtSearch.setPlaceholderText("name…")
        fb.addWidget(self._txtSearch, 1)
        lv.addLayout(fb)

        self._tblItems = _tbl(["Item", "Category", "UoM", "Stock", "Status"])
        self._tblItems.setColumnWidth(0, 180)
        self._tblItems.setColumnWidth(1, 80)
        self._tblItems.setColumnWidth(2, 50)
        self._tblItems.setColumnWidth(3, 60)
        self._tblItems.setSortingEnabled(False)
        lv.addWidget(self._tblItems, 1)
        left.setMinimumWidth(360)

        # ── Right: detail + lots ─────────────────────────────────────────────
        right = QWidget()
        rv    = QVBoxLayout(right)
        rv.setContentsMargins(4, 0, 0, 0)

        # action buttons
        ab = QHBoxLayout()
        self._btnNew    = QPushButton("New");    self._btnNew.setStyleSheet(_BTN_ADD_SS)
        self._btnEdit   = QPushButton("Edit");   self._btnEdit.setStyleSheet(_BTN_SS)
        self._btnSave   = QPushButton("Save");   self._btnSave.setStyleSheet(_BTN_ADD_SS)
        self._btnCancel = QPushButton("Cancel"); self._btnCancel.setStyleSheet(_BTN_SS)
        self._btnDelete = QPushButton("Delete"); self._btnDelete.setStyleSheet(_BTN_DEL_SS)
        for b in [self._btnNew, self._btnEdit, self._btnSave, self._btnCancel, self._btnDelete]:
            ab.addWidget(b)
        ab.addStretch(1)
        ab.addWidget(make_help_button(self, "cims"))
        rv.addLayout(ab)

        # detail form
        g = QGroupBox("Item Details")
        gf = QFormLayout(g); gf.setSpacing(5)

        self._txtName   = QLineEdit()
        self._cmbCatDet = QComboBox()
        for c in _CATEGORIES: self._cmbCatDet.addItem(c, c)
        self._txtCAS    = QLineEdit(); self._txtCAS.setMaximumWidth(120)
        self._txtUnit   = QLineEdit(); self._txtUnit.setMaximumWidth(80)
        self._spnReord  = QDoubleSpinBox(); self._spnReord.setRange(0, 1e9); self._spnReord.setDecimals(3)
        self._spnReordQ = QDoubleSpinBox(); self._spnReordQ.setRange(0, 1e9); self._spnReordQ.setDecimals(3)
        self._txtLoc    = QLineEdit()
        self._chkActive = QCheckBox("Active"); self._chkActive.setChecked(True)
        self._txtRem    = QTextEdit(); self._txtRem.setFixedHeight(54)

        gf.addRow("Name:",          self._txtName)
        gf.addRow("Category:",      self._cmbCatDet)

        row_cas = QHBoxLayout()
        row_cas.addWidget(self._txtCAS)
        row_cas.addWidget(QLabel("  Unit:"))
        row_cas.addWidget(self._txtUnit)
        row_cas.addStretch(1)
        gf.addRow("CAS #:",         row_cas)

        row_ro = QHBoxLayout()
        row_ro.addWidget(self._spnReord)
        row_ro.addWidget(QLabel("  Order qty:"))
        row_ro.addWidget(self._spnReordQ)
        row_ro.addStretch(1)
        gf.addRow("Reorder at:",    row_ro)

        gf.addRow("Storage:",       self._txtLoc)
        gf.addRow("",               self._chkActive)
        gf.addRow("Remarks:",       self._txtRem)
        rv.addWidget(g)

        # lots sub-table
        lg = QGroupBox("Lots in Stock")
        lgl = QVBoxLayout(lg)

        lot_actions = QHBoxLayout()
        self._btnAddLot    = QPushButton("Add Lot");      self._btnAddLot.setStyleSheet(_BTN_ADD_SS)
        self._btnObsolete  = QPushButton("Mark Obsolete"); self._btnObsolete.setStyleSheet(_BTN_SS)
        self._btnUsage     = QPushButton("Record Usage");  self._btnUsage.setStyleSheet(_BTN_SS)
        for b in [self._btnAddLot, self._btnObsolete, self._btnUsage]:
            lot_actions.addWidget(b)
        lot_actions.addStretch(1)
        lgl.addLayout(lot_actions)

        self._tblLots = _tbl(["Lot #", "Supplier", "Received", "Expiry", "Qty Rem.", "Unit Cost", "Obsolete"])
        self._tblLots.setColumnWidth(0, 90)
        self._tblLots.setColumnWidth(1, 110)
        self._tblLots.setColumnWidth(2, 80)
        self._tblLots.setColumnWidth(3, 80)
        self._tblLots.setColumnWidth(4, 70)
        self._tblLots.setColumnWidth(5, 80)
        self._tblLots.setSortingEnabled(False)
        self._tblLots.setFixedHeight(160)
        lgl.addWidget(self._tblLots)
        rv.addWidget(lg)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        QVBoxLayout(self).addWidget(splitter)

        # signals
        self._cmbCat.currentIndexChanged.connect(self._load_items)
        self._txtSearch.textChanged.connect(self._load_items)
        self._tblItems.itemSelectionChanged.connect(self._on_item_sel)
        self._btnNew.clicked.connect(self._on_new)
        self._btnEdit.clicked.connect(self._on_edit)
        self._btnSave.clicked.connect(self._on_save)
        self._btnCancel.clicked.connect(self._on_cancel)
        self._btnDelete.clicked.connect(self._on_delete)
        self._btnAddLot.clicked.connect(self._on_add_lot)
        self._btnObsolete.clicked.connect(self._on_obsolete_lot)
        self._btnUsage.clicked.connect(self._on_record_usage)

        self._set_ro(True)
        self._load_suppliers()
        self._load_items()

    # ── data helpers ─────────────────────────────────────────────────────────

    def _load_suppliers(self):
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text(
                    "SELECT supplier_id, supplier_name FROM cims_supplier "
                    "WHERE is_active=1 ORDER BY supplier_name"
                )).fetchall()
            self._suppliers = [(r.supplier_id, r.supplier_name) for r in rows]
        except Exception:
            self._suppliers = []

    def _load_items(self):
        cat    = self._cmbCat.currentData()
        search = self._txtSearch.text().strip()
        try:
            with db_manager.get_connection() as conn:
                sql = """
                    SELECT i.item_id, i.item_name, i.category, i.unit_of_measure,
                           i.reorder_point,
                           COALESCE(SUM(l.qty_remaining),0) AS stock
                    FROM cims_item i
                    LEFT JOIN cims_lot l ON l.item_id=i.item_id AND l.is_obsolete=0
                    WHERE i.is_active=1
                """
                p: dict = {}
                if cat:
                    sql += " AND i.category=:cat"; p["cat"] = cat
                if search:
                    sql += " AND LOWER(i.item_name) LIKE :q"; p["q"] = f"%{search.lower()}%"
                sql += " GROUP BY i.item_id, i.item_name, i.category, i.unit_of_measure, i.reorder_point ORDER BY i.item_name"
                rows = conn.execute(text(sql), p).fetchall()
        except Exception as e:
            log.error("CIMS load_items: %s", e)
            return

        self._tblItems.setSortingEnabled(False)
        self._tblItems.setRowCount(0)
        for row in rows:
            r   = self._tblItems.rowCount(); self._tblItems.insertRow(r)
            stk = float(row.stock or 0)
            ro  = float(row.reorder_point or 0)
            if stk <= 0:
                status, bg = "Empty", _RED
            elif ro > 0 and stk <= ro:
                status, bg = "Low", _AMBER
            elif ro > 0 and stk <= ro * 1.2:
                status, bg = "Near", QColor("#FFF8E1")
            else:
                status, bg = "OK", QColor("#FFFFFF")

            cells = [row.item_name, row.category, row.unit_of_measure,
                     f"{stk:g}", status]
            for c, val in enumerate(cells):
                it = _cell(val)
                it.setBackground(bg)
                if c == 0:
                    it.setData(Qt.UserRole, row.item_id)
                self._tblItems.setItem(r, c, it)

    def _on_item_sel(self):
        rows = self._tblItems.selectedItems()
        if not rows:
            self._current_item_id = None
            self._clear_detail()
            self._set_ro(True)
            return
        item_id = self._tblItems.item(rows[0].row(), 0).data(Qt.UserRole)
        self._current_item_id = item_id
        self._load_detail(item_id)
        self._load_lots(item_id)
        self._set_ro(True)

    def _load_detail(self, item_id: int):
        try:
            with db_manager.get_connection() as conn:
                row = conn.execute(text(
                    "SELECT * FROM cims_item WHERE item_id=:id"
                ), {"id": item_id}).fetchone()
            if not row:
                return
            self._txtName.setText(row.item_name or "")
            self._cmbCatDet.setCurrentIndex(self._cmbCatDet.findData(row.category))
            self._txtCAS.setText(row.cas_number or "")
            self._txtUnit.setText(row.unit_of_measure or "")
            self._spnReord.setValue(float(row.reorder_point or 0))
            self._spnReordQ.setValue(float(row.reorder_qty or 0))
            self._txtLoc.setText(row.storage_location or "")
            self._chkActive.setChecked(bool(row.is_active))
            self._txtRem.setPlainText(row.remarks or "")
        except Exception as e:
            log.error("CIMS load_detail: %s", e)

    def _load_lots(self, item_id: Optional[int]):
        self._tblLots.setRowCount(0)
        if item_id is None:
            return
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text("""
                    SELECT l.lot_id, l.lot_number, s.supplier_name,
                           l.date_received, l.expiry_date,
                           l.qty_remaining, l.unit_cost, l.is_obsolete
                    FROM cims_lot l
                    LEFT JOIN cims_supplier s ON s.supplier_id=l.supplier_id
                    WHERE l.item_id=:id
                    ORDER BY l.date_received DESC, l.lot_id DESC
                """), {"id": item_id}).fetchall()
        except Exception as e:
            log.error("CIMS load_lots: %s", e); return

        today = date.today()
        warn_days = int(get_global_value("cims_expiry_warn_days", 30))

        for row in rows:
            r = self._tblLots.rowCount(); self._tblLots.insertRow(r)
            obs = bool(row.is_obsolete)
            exp = row.expiry_date
            if obs:
                bg = _GREY
            elif exp and exp < today:
                bg = _RED
            elif exp and exp <= today + timedelta(days=warn_days):
                bg = _AMBER
            else:
                bg = QColor("#FFFFFF")

            cells = [
                row.lot_number or "—",
                row.supplier_name or "—",
                str(row.date_received) if row.date_received else "—",
                str(exp) if exp else "—",
                f"{float(row.qty_remaining or 0):g}",
                f"{float(row.unit_cost):.4f}" if row.unit_cost else "—",
                "Yes" if obs else "",
            ]
            font = QFont(); font.setItalic(obs)
            for c, val in enumerate(cells):
                it = _cell(val)
                it.setBackground(bg)
                it.setFont(font)
                if c == 0:
                    it.setData(Qt.UserRole, row.lot_id)
                    it.setData(Qt.UserRole + 1, float(row.qty_remaining or 0))
                self._tblLots.setItem(r, c, it)

    # ── form state ────────────────────────────────────────────────────────────

    def _set_ro(self, ro: bool):
        editable = [self._txtName, self._txtCAS, self._txtUnit, self._txtLoc]
        for w in editable:
            w.setReadOnly(ro)
        for w in [self._cmbCatDet, self._chkActive, self._spnReord, self._spnReordQ]:
            w.setEnabled(not ro)
        self._txtRem.setReadOnly(ro)
        self._btnSave.setEnabled(not ro)
        self._btnCancel.setEnabled(not ro)
        has = self._current_item_id is not None
        self._btnEdit.setEnabled(ro and has)
        self._btnDelete.setEnabled(ro and has)
        self._btnNew.setEnabled(ro)
        self._btnAddLot.setEnabled(ro and has)
        self._btnObsolete.setEnabled(ro and has)
        self._btnUsage.setEnabled(ro and has)
        self._tblItems.setEnabled(ro)

    def _clear_detail(self):
        self._txtName.clear(); self._txtCAS.clear()
        self._txtUnit.clear(); self._txtLoc.clear()
        self._spnReord.setValue(0); self._spnReordQ.setValue(0)
        self._chkActive.setChecked(True); self._txtRem.clear()
        self._tblLots.setRowCount(0)

    # ── CRUD slots ────────────────────────────────────────────────────────────

    def _on_new(self):
        self._is_new = True
        self._current_item_id = None
        self._clear_detail()
        self._set_ro(False)
        self._txtName.setFocus()

    def _on_edit(self):
        self._is_new = False
        self._set_ro(False)

    def _on_cancel(self):
        self._is_new = False
        self._set_ro(True)
        if self._current_item_id:
            self._load_detail(self._current_item_id)

    def _on_save(self):
        name = self._txtName.text().strip()
        if not name:
            show_message(self, "Validation", "Item name is required.")
            return
        unit = self._txtUnit.text().strip()
        if not unit:
            show_message(self, "Validation", "Unit of measure is required.")
            return
        p = {
            "name":  name,
            "cat":   self._cmbCatDet.currentData(),
            "cas":   self._txtCAS.text().strip() or None,
            "unit":  unit,
            "ro":    self._spnReord.value(),
            "roq":   self._spnReordQ.value() or None,
            "loc":   self._txtLoc.text().strip() or None,
            "act":   1 if self._chkActive.isChecked() else 0,
            "rem":   self._txtRem.toPlainText().strip() or None,
            "now":   datetime.now(),
            "user":  getpass.getuser(),
        }
        try:
            with db_manager.get_connection() as conn:
                if self._is_new:
                    res = conn.execute(text("""
                        INSERT INTO cims_item
                          (item_name,category,cas_number,unit_of_measure,reorder_point,
                           reorder_qty,storage_location,is_active,remarks,
                           createdatestamp,createuserstamp,modifdatestamp,modifuserstamp)
                        VALUES
                          (:name,:cat,:cas,:unit,:ro,:roq,:loc,:act,:rem,
                           :now,:user,:now,:user)
                        RETURNING item_id
                    """), p).fetchone()
                    self._current_item_id = res[0]
                else:
                    p["id"] = self._current_item_id
                    conn.execute(text("""
                        UPDATE cims_item SET
                          item_name=:name, category=:cat, cas_number=:cas,
                          unit_of_measure=:unit, reorder_point=:ro, reorder_qty=:roq,
                          storage_location=:loc, is_active=:act, remarks=:rem,
                          modifdatestamp=:now, modifuserstamp=:user
                        WHERE item_id=:id
                    """), p)
                conn.commit()
            self._is_new = False
            self._set_ro(True)
            self._load_items()
            self._load_lots(self._current_item_id)
        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))

    def _on_delete(self):
        if QMessageBox.question(
            self, "Delete Item",
            "Delete this item?\n\nThis will fail if the item has existing lots.",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return
        try:
            with db_manager.get_connection() as conn:
                conn.execute(text(
                    "DELETE FROM cims_item WHERE item_id=:id"
                ), {"id": self._current_item_id})
                conn.commit()
            self._current_item_id = None
            self._clear_detail()
            self._load_items()
            self._set_ro(True)
        except Exception as e:
            QMessageBox.critical(self, "Delete Error", str(e))

    # ── lot slots ─────────────────────────────────────────────────────────────

    def _on_add_lot(self):
        if not self._current_item_id:
            return
        item_name = self._txtName.text()
        dlg = _LotDialog(self._current_item_id, item_name, self._suppliers, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        v = dlg.values()
        v["now"] = datetime.now(); v["user"] = getpass.getuser()
        try:
            with db_manager.get_connection() as conn:
                conn.execute(text("""
                    INSERT INTO cims_lot
                      (item_id,supplier_id,lot_number,date_received,expiry_date,
                       qty_received,qty_remaining,unit_cost,invoice_ref,remarks,
                       createdatestamp,createuserstamp,modifdatestamp,modifuserstamp)
                    VALUES
                      (:item_id,:supplier_id,:lot_number,:date_received,:expiry_date,
                       :qty_received,:qty_remaining,:unit_cost,:invoice_ref,:remarks,
                       :now,:user,:now,:user)
                """), v)
                conn.commit()
            self._load_lots(self._current_item_id)
            self._load_items()
        except Exception as e:
            QMessageBox.critical(self, "Add Lot Error", str(e))

    def _selected_lot(self) -> tuple[Optional[int], float, str]:
        """Return (lot_id, qty_remaining, lot_label) for the selected lot row."""
        sel = self._tblLots.selectedItems()
        if not sel:
            return None, 0.0, ""
        row = sel[0].row()
        cell0 = self._tblLots.item(row, 0)
        lot_id  = cell0.data(Qt.UserRole)
        qty_rem = cell0.data(Qt.UserRole + 1)
        label   = cell0.text()
        return lot_id, qty_rem, label

    def _on_obsolete_lot(self):
        lot_id, _, label = self._selected_lot()
        if lot_id is None:
            show_message(self, "Mark Obsolete", "Select a lot first.")
            return
        if QMessageBox.question(
            self, "Mark Obsolete",
            f"Mark lot '{label}' as obsolete and zero its remaining quantity?",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return
        try:
            with db_manager.get_connection() as conn:
                conn.execute(text("""
                    UPDATE cims_lot SET is_obsolete=1, qty_remaining=0,
                      modifdatestamp=:now, modifuserstamp=:user
                    WHERE lot_id=:id
                """), {"id": lot_id, "now": datetime.now(), "user": getpass.getuser()})
                conn.commit()
            self._load_lots(self._current_item_id)
            self._load_items()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _on_record_usage(self):
        lot_id, qty_rem, label = self._selected_lot()
        if lot_id is None:
            show_message(self, "Record Usage", "Select a lot first.")
            return
        dlg = _UsageDialog(lot_id, label, qty_rem, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        v = dlg.values(); v["now"] = datetime.now(); v["user"] = getpass.getuser()
        qty = v["quantity"]
        move = v["movement_type"]
        try:
            with db_manager.get_connection() as conn:
                conn.execute(text("""
                    INSERT INTO cims_usage
                      (lot_id,movement_type,quantity,movement_date,used_by,
                       run_module,run_id,purpose,createdatestamp,createuserstamp)
                    VALUES
                      (:lot_id,:movement_type,:quantity,:now,:used_by,
                       :run_module,:run_id,:purpose,:now,:user)
                """), v)
                # decrement qty_remaining for USED and SCRAPPED
                if move in ("USED", "SCRAPPED"):
                    new_q = max(0, qty_rem - qty)
                    conn.execute(text("""
                        UPDATE cims_lot SET qty_remaining=:q,
                          modifdatestamp=:now, modifuserstamp=:user
                        WHERE lot_id=:id
                    """), {"q": new_q, "id": lot_id, "now": v["now"], "user": v["user"]})
                elif move == "ADJUSTED":
                    conn.execute(text("""
                        UPDATE cims_lot SET qty_remaining=:q,
                          modifdatestamp=:now, modifuserstamp=:user
                        WHERE lot_id=:id
                    """), {"q": qty, "id": lot_id, "now": v["now"], "user": v["user"]})
                conn.commit()
            self._load_lots(self._current_item_id)
            self._load_items()
        except Exception as e:
            QMessageBox.critical(self, "Usage Error", str(e))

    def get_alert_counts(self) -> dict:
        """Return low_stock and expiring_soon counts for sidebar badge."""
        warn_days = int(get_global_value("cims_expiry_warn_days", 30))
        today     = date.today()
        counts    = {"low_stock": 0, "expiring_soon": 0}
        try:
            with db_manager.get_connection() as conn:
                # low stock
                rows = conn.execute(text("""
                    SELECT COUNT(*) AS n FROM (
                      SELECT i.item_id
                      FROM cims_item i
                      LEFT JOIN cims_lot l ON l.item_id=i.item_id AND l.is_obsolete=0
                      WHERE i.is_active=1 AND i.reorder_point > 0
                      GROUP BY i.item_id, i.reorder_point
                      HAVING COALESCE(SUM(l.qty_remaining),0) <= i.reorder_point
                    ) s
                """)).fetchone()
                counts["low_stock"] = int(rows[0] or 0)
                # expiring soon
                rows2 = conn.execute(text("""
                    SELECT COUNT(*) AS n FROM cims_lot
                    WHERE is_obsolete=0
                      AND expiry_date IS NOT NULL
                      AND expiry_date <= :warn
                """), {"warn": today + timedelta(days=warn_days)}).fetchone()
                counts["expiring_soon"] = int(rows2[0] or 0)
        except Exception:
            pass
        return counts


# ── Usage Log panel ───────────────────────────────────────────────────────────

class CIMSUsagePanel(QWidget):
    """Read-only movement history with filters and CSV export."""

    def __init__(self, parent=None):
        super().__init__(parent)
        v = QVBoxLayout(self)

        # filter bar
        fb = QHBoxLayout()
        fb.addWidget(QLabel("Item:"))
        self._cmbItem = QComboBox(); self._cmbItem.setMinimumWidth(160)
        fb.addWidget(self._cmbItem)
        fb.addWidget(QLabel("Module:"))
        self._cmbMod = QComboBox()
        self._cmbMod.addItem("All", None)
        for m in _MODULES: self._cmbMod.addItem(m, m)
        fb.addWidget(self._cmbMod)
        fb.addWidget(QLabel("Type:"))
        self._cmbType = QComboBox()
        self._cmbType.addItem("All", None)
        for t in _MOVE_TYPES: self._cmbType.addItem(t, t)
        fb.addWidget(self._cmbType)
        fb.addWidget(QLabel("From:"))
        self._dtFrom = QDateEdit(QDate.currentDate().addDays(-90))
        self._dtFrom.setCalendarPopup(True)
        fb.addWidget(self._dtFrom)
        fb.addWidget(QLabel("To:"))
        self._dtTo = QDateEdit(QDate.currentDate())
        self._dtTo.setCalendarPopup(True)
        fb.addWidget(self._dtTo)
        self._btnApply  = QPushButton("Apply");      self._btnApply.setStyleSheet(_BTN_SS)
        self._btnReset  = QPushButton("Reset");      self._btnReset.setStyleSheet(_BTN_SS)
        self._btnExport = QPushButton("Export CSV"); self._btnExport.setStyleSheet(_BTN_SS)
        fb.addWidget(self._btnApply); fb.addWidget(self._btnReset); fb.addWidget(self._btnExport)
        v.addLayout(fb)

        self._tblLog = _tbl(["Date", "Item", "Lot #", "Type", "Qty", "Used By", "Module/Run", "Purpose"])
        self._tblLog.setColumnWidth(0, 140)
        self._tblLog.setColumnWidth(1, 160)
        self._tblLog.setColumnWidth(2, 80)
        self._tblLog.setColumnWidth(3, 80)
        self._tblLog.setColumnWidth(4, 60)
        self._tblLog.setColumnWidth(5, 90)
        self._tblLog.setColumnWidth(6, 90)
        v.addWidget(self._tblLog, 1)

        self._btnApply.clicked.connect(self._load)
        self._btnReset.clicked.connect(self._reset)
        self._btnExport.clicked.connect(self._export_csv)

        self._load_item_combo()
        self._load()

    def _load_item_combo(self):
        self._cmbItem.clear()
        self._cmbItem.addItem("All items", None)
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text(
                    "SELECT item_id, item_name FROM cims_item ORDER BY item_name"
                )).fetchall()
            for r in rows:
                self._cmbItem.addItem(r.item_name, r.item_id)
        except Exception:
            pass

    def _reset(self):
        self._cmbItem.setCurrentIndex(0)
        self._cmbMod.setCurrentIndex(0)
        self._cmbType.setCurrentIndex(0)
        self._dtFrom.setDate(QDate.currentDate().addDays(-90))
        self._dtTo.setDate(QDate.currentDate())
        self._load()

    def _load(self):
        item_id = self._cmbItem.currentData()
        mod     = self._cmbMod.currentData()
        mtype   = self._cmbType.currentData()
        df      = self._dtFrom.date(); dt = self._dtTo.date()
        date_from = date(df.year(), df.month(), df.day())
        date_to   = date(dt.year(), dt.month(), dt.day())

        sql = """
            SELECT u.movement_date, i.item_name, l.lot_number,
                   u.movement_type, u.quantity, u.used_by,
                   u.run_module, u.run_id, u.purpose
            FROM cims_usage u
            JOIN cims_lot l  ON l.lot_id  = u.lot_id
            JOIN cims_item i ON i.item_id = l.item_id
            WHERE u.movement_date >= :df AND u.movement_date < :dt
        """
        p: dict = {"df": date_from, "dt": date_to}
        if item_id:
            sql += " AND i.item_id=:iid"; p["iid"] = item_id
        if mod:
            sql += " AND u.run_module=:mod"; p["mod"] = mod
        if mtype:
            sql += " AND u.movement_type=:mt"; p["mt"] = mtype
        sql += " ORDER BY u.movement_date DESC"

        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text(sql), p).fetchall()
        except Exception as e:
            log.error("CIMS usage log: %s", e); return

        self._tblLog.setSortingEnabled(False)
        self._tblLog.setRowCount(0)
        for row in rows:
            r = self._tblLog.rowCount(); self._tblLog.insertRow(r)
            bg   = _MOVE_COLOURS.get(row.movement_type, QColor("#FFFFFF"))
            run  = f"{row.run_module}/{row.run_id}" if row.run_module and row.run_id else (row.run_module or "")
            vals = [
                str(row.movement_date)[:19] if row.movement_date else "",
                row.item_name or "",
                row.lot_number or "—",
                row.movement_type or "",
                f"{float(row.quantity or 0):g}",
                row.used_by or "",
                run,
                row.purpose or "",
            ]
            for c, val in enumerate(vals):
                it = _cell(val)
                it.setBackground(bg)
                self._tblLog.setItem(r, c, it)
        self._tblLog.setSortingEnabled(True)

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Usage Log", "cims_usage_log.csv", "CSV (*.csv)"
        )
        if not path:
            return
        try:
            headers = ["Date", "Item", "Lot #", "Type", "Qty", "Used By", "Module/Run", "Purpose"]
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for r in range(self._tblLog.rowCount()):
                    w.writerow([
                        (self._tblLog.item(r, c).text() if self._tblLog.item(r, c) else "")
                        for c in range(self._tblLog.columnCount())
                    ])
            show_message(self, "Export", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ── Reports panel ─────────────────────────────────────────────────────────────

class CIMSReportPanel(QWidget):
    """Stock valuation and usage-per-run reports with Excel and PDF export."""

    def __init__(self, parent=None):
        super().__init__(parent)
        v = QVBoxLayout(self)

        # ── Stock Valuation ──────────────────────────────────────────────────
        g1 = QGroupBox("Stock Valuation")
        g1l = QVBoxLayout(g1)
        b1 = QHBoxLayout()
        self._btnGenVal    = QPushButton("Generate");     self._btnGenVal.setStyleSheet(_BTN_SS)
        self._btnExcelVal  = QPushButton("Export Excel"); self._btnExcelVal.setStyleSheet(_BTN_SS)
        self._btnPDFVal    = QPushButton("Export PDF");   self._btnPDFVal.setStyleSheet(_BTN_SS)
        self._btnExcelVal.setEnabled(False); self._btnPDFVal.setEnabled(False)
        b1.addWidget(self._btnGenVal); b1.addWidget(self._btnExcelVal); b1.addWidget(self._btnPDFVal)
        b1.addStretch(1); g1l.addLayout(b1)

        self._tblVal = _tbl(["Category", "Item", "UoM", "Stock Qty", "Unit Cost", "Total Value", "Reorder?"])
        self._tblVal.setColumnWidth(0, 90); self._tblVal.setColumnWidth(1, 180)
        self._tblVal.setColumnWidth(2, 55); self._tblVal.setColumnWidth(3, 80)
        self._tblVal.setColumnWidth(4, 80); self._tblVal.setColumnWidth(5, 90)
        g1l.addWidget(self._tblVal)
        v.addWidget(g1)

        # ── Usage per Run ────────────────────────────────────────────────────
        g2 = QGroupBox("Usage per Run Type")
        g2l = QVBoxLayout(g2)
        b2  = QHBoxLayout()
        b2.addWidget(QLabel("From:"))
        self._dtFrom2 = QDateEdit(QDate.currentDate().addDays(-365))
        self._dtFrom2.setCalendarPopup(True); b2.addWidget(self._dtFrom2)
        b2.addWidget(QLabel("To:"))
        self._dtTo2 = QDateEdit(QDate.currentDate())
        self._dtTo2.setCalendarPopup(True); b2.addWidget(self._dtTo2)
        b2.addWidget(QLabel("Module:"))
        self._cmbRunMod = QComboBox()
        self._cmbRunMod.addItem("All", None)
        for m in _MODULES: self._cmbRunMod.addItem(m, m)
        b2.addWidget(self._cmbRunMod)
        self._btnGenUse    = QPushButton("Generate");     self._btnGenUse.setStyleSheet(_BTN_SS)
        self._btnExcelUse  = QPushButton("Export Excel"); self._btnExcelUse.setStyleSheet(_BTN_SS)
        self._btnExcelUse.setEnabled(False)
        b2.addWidget(self._btnGenUse); b2.addWidget(self._btnExcelUse)
        b2.addStretch(1); g2l.addLayout(b2)

        self._tblUse = _tbl(["Module", "Run ID", "Item", "Lot #", "Qty Used", "Purpose"])
        self._tblUse.setColumnWidth(0, 70); self._tblUse.setColumnWidth(1, 70)
        self._tblUse.setColumnWidth(2, 180); self._tblUse.setColumnWidth(3, 80)
        self._tblUse.setColumnWidth(4, 80)
        g2l.addWidget(self._tblUse)
        v.addWidget(g2)

        self._valuation_data: list = []
        self._usage_data:     list = []

        self._btnGenVal.clicked.connect(self._gen_valuation)
        self._btnExcelVal.clicked.connect(self._excel_valuation)
        self._btnPDFVal.clicked.connect(self._pdf_valuation)
        self._btnGenUse.clicked.connect(self._gen_usage)
        self._btnExcelUse.clicked.connect(self._excel_usage)

    # ── valuation ─────────────────────────────────────────────────────────────

    def _gen_valuation(self):
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text("""
                    SELECT i.category, i.item_name, i.unit_of_measure,
                           i.reorder_point,
                           COALESCE(SUM(l.qty_remaining), 0)   AS stock,
                           AVG(CASE WHEN l.unit_cost>0 THEN l.unit_cost END) AS avg_cost
                    FROM cims_item i
                    LEFT JOIN cims_lot l ON l.item_id=i.item_id AND l.is_obsolete=0
                    WHERE i.is_active=1
                    GROUP BY i.item_id, i.category, i.item_name, i.unit_of_measure, i.reorder_point
                    ORDER BY i.category, i.item_name
                """)).fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e)); return

        self._valuation_data = []
        self._tblVal.setRowCount(0)
        for row in rows:
            stk  = float(row.stock or 0)
            cost = float(row.avg_cost or 0)
            total = stk * cost
            below = stk <= float(row.reorder_point or 0) and float(row.reorder_point or 0) > 0
            self._valuation_data.append({
                "category": row.category, "item": row.item_name,
                "uom": row.unit_of_measure, "stock": stk,
                "cost": cost, "total": total, "below": below,
            })
            r = self._tblVal.rowCount(); self._tblVal.insertRow(r)
            cells = [
                row.category, row.item_name, row.unit_of_measure,
                f"{stk:g}", f"{cost:.4f}" if cost else "—",
                f"{total:.2f}" if cost else "—",
                "⚠ Low" if below else "",
            ]
            for c, val in enumerate(cells):
                it = _cell(val)
                if below and c == 6:
                    it.setForeground(QColor("#C62828"))
                    it.setFont(QFont("", -1, QFont.Bold))
                self._tblVal.setItem(r, c, it)

        self._btnExcelVal.setEnabled(True)
        self._btnPDFVal.setEnabled(True)

    def _excel_valuation(self):
        if not self._valuation_data:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Stock Valuation", "cims_stock_valuation.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "Stock Valuation"
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="37474F")
            headers  = ["Category", "Item", "UoM", "Stock Qty", "Unit Cost", "Total Value", "Reorder?"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(1, ci, h); c.font = hdr_font; c.fill = hdr_fill
            for ri, d in enumerate(self._valuation_data, 2):
                ws.cell(ri, 1, d["category"])
                ws.cell(ri, 2, d["item"])
                ws.cell(ri, 3, d["uom"])
                ws.cell(ri, 4, d["stock"])
                ws.cell(ri, 5, d["cost"] if d["cost"] else None)
                ws.cell(ri, 6, round(d["total"], 2) if d["cost"] else None)
                c7 = ws.cell(ri, 7, "Low" if d["below"] else "")
                if d["below"]:
                    c7.font = Font(bold=True, color="C62828")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 3, 45
                )
            wb.save(path)
            show_message(self, "Export", f"Saved:\n{path}")
        except ImportError:
            show_message(self, "Export", "openpyxl is required: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _pdf_valuation(self):
        if not self._valuation_data:
            return
        rows_html = ""
        for d in self._valuation_data:
            warn_style = ' style="color:#C62828;font-weight:bold"' if d["below"] else ""
            rows_html += (
                f"<tr><td>{d['category']}</td><td>{d['item']}</td>"
                f"<td>{d['uom']}</td><td align='right'>{d['stock']:g}</td>"
                f"<td align='right'>{'%.4f'%d['cost'] if d['cost'] else '—'}</td>"
                f"<td align='right'>{'%.2f'%d['total'] if d['cost'] else '—'}</td>"
                f"<td{warn_style}>{'⚠ Low' if d['below'] else ''}</td></tr>\n"
            )
        html = f"""<!DOCTYPE html><html><head><meta charset='utf-8'>
        <style>
          body{{font-family:Arial;font-size:11px}}
          h2{{color:#37474F}}
          table{{border-collapse:collapse;width:100%}}
          th{{background:#37474F;color:white;padding:4px 6px;text-align:left}}
          td{{padding:3px 6px;border-bottom:1px solid #e0e0e0}}
          tr:nth-child(even) td{{background:#f5f5f5}}
        </style></head><body>
        <h2>IsoWorks CIMS — Stock Valuation Report</h2>
        <p>Generated: {datetime.now():%Y-%m-%d %H:%M}</p>
        <table><tr>
          <th>Category</th><th>Item</th><th>UoM</th>
          <th>Stock Qty</th><th>Unit Cost</th><th>Total Value</th><th>Reorder?</th>
        </tr>{rows_html}</table></body></html>"""
        printer = QPrinter(QPrinter.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec_():
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print_(printer)

    # ── usage per run ─────────────────────────────────────────────────────────

    def _gen_usage(self):
        df  = self._dtFrom2.date(); dt = self._dtTo2.date()
        d_from = date(df.year(), df.month(), df.day())
        d_to   = date(dt.year(), dt.month(), dt.day())
        mod    = self._cmbRunMod.currentData()
        sql = """
            SELECT u.run_module, u.run_id, i.item_name, l.lot_number,
                   u.quantity, u.purpose
            FROM cims_usage u
            JOIN cims_lot  l ON l.lot_id  = u.lot_id
            JOIN cims_item i ON i.item_id = l.item_id
            WHERE u.movement_type='USED'
              AND u.movement_date >= :df AND u.movement_date < :dt
              AND u.run_module IS NOT NULL
        """
        p: dict = {"df": d_from, "dt": d_to}
        if mod:
            sql += " AND u.run_module=:mod"; p["mod"] = mod
        sql += " ORDER BY u.run_module, u.run_id, i.item_name"
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text(sql), p).fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e)); return

        self._usage_data = []
        self._tblUse.setRowCount(0)
        for row in rows:
            self._usage_data.append({
                "module": row.run_module, "run_id": row.run_id,
                "item": row.item_name, "lot": row.lot_number,
                "qty": float(row.quantity or 0), "purpose": row.purpose,
            })
            r = self._tblUse.rowCount(); self._tblUse.insertRow(r)
            for c, val in enumerate([
                row.run_module or "", str(row.run_id) if row.run_id else "",
                row.item_name, row.lot_number or "—",
                f"{float(row.quantity or 0):g}", row.purpose or "",
            ]):
                self._tblUse.setItem(r, c, _cell(val))
        self._btnExcelUse.setEnabled(bool(self._usage_data))

    def _excel_usage(self):
        if not self._usage_data:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Usage Report", "cims_usage_per_run.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Usage per Run"
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="37474F")
            headers  = ["Module", "Run ID", "Item", "Lot #", "Qty Used", "Purpose"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(1, ci, h); c.font = hdr_font; c.fill = hdr_fill
            for ri, d in enumerate(self._usage_data, 2):
                ws.cell(ri, 1, d["module"]); ws.cell(ri, 2, d["run_id"])
                ws.cell(ri, 3, d["item"]);   ws.cell(ri, 4, d["lot"])
                ws.cell(ri, 5, d["qty"]);    ws.cell(ri, 6, d["purpose"])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 3, 50
                )
            wb.save(path)
            show_message(self, "Export", f"Saved:\n{path}")
        except ImportError:
            show_message(self, "Export", "openpyxl is required: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ── Supplier management (lightweight, embedded in a dialog) ───────────────────

class _SupplierManagementDialog(QDialog):
    """Simple CRUD for cims_supplier — opened from a toolbar button."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Supplier Management")
        self.resize(700, 420)
        self._current_id: Optional[int] = None
        self._is_new = False

        v = QVBoxLayout(self)

        # action bar
        ab = QHBoxLayout()
        self._btnNew    = QPushButton("New");    self._btnNew.setStyleSheet(_BTN_ADD_SS)
        self._btnEdit   = QPushButton("Edit");   self._btnEdit.setStyleSheet(_BTN_SS)
        self._btnSave   = QPushButton("Save");   self._btnSave.setStyleSheet(_BTN_ADD_SS)
        self._btnCancel = QPushButton("Cancel"); self._btnCancel.setStyleSheet(_BTN_SS)
        self._btnDel    = QPushButton("Delete"); self._btnDel.setStyleSheet(_BTN_DEL_SS)
        for b in [self._btnNew, self._btnEdit, self._btnSave, self._btnCancel, self._btnDel]:
            ab.addWidget(b)
        ab.addStretch(1); v.addLayout(ab)

        spl = QSplitter(Qt.Horizontal)

        # list
        self._tbl = _tbl(["ID", "Name", "Category", "Contact", "Active"])
        self._tbl.setColumnWidth(0, 40); self._tbl.setColumnWidth(1, 160)
        self._tbl.setColumnWidth(2, 80); self._tbl.setColumnWidth(3, 120)
        spl.addWidget(self._tbl)

        # detail
        det = QWidget()
        f = QFormLayout(det); f.setSpacing(6)
        self._txtName    = QLineEdit()
        self._cmbCat     = QComboBox()
        for c in ["Chemical", "Gas", "LabSupply", "Instrument", "General"]:
            self._cmbCat.addItem(c, c)
        self._txtContact = QLineEdit()
        self._txtPhone   = QLineEdit()
        self._txtEmail   = QLineEdit()
        self._txtWeb     = QLineEdit()
        self._chkActive  = QCheckBox("Active"); self._chkActive.setChecked(True)
        self._txtRem     = QTextEdit(); self._txtRem.setFixedHeight(54)

        f.addRow("Name:",     self._txtName)
        f.addRow("Category:", self._cmbCat)
        f.addRow("Contact:",  self._txtContact)
        f.addRow("Phone:",    self._txtPhone)
        f.addRow("Email:",    self._txtEmail)
        f.addRow("Website:",  self._txtWeb)
        f.addRow("",          self._chkActive)
        f.addRow("Remarks:",  self._txtRem)
        spl.addWidget(det)
        v.addWidget(spl, 1)

        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(self.accept)
        v.addWidget(bb)

        self._btnNew.clicked.connect(self._on_new)
        self._btnEdit.clicked.connect(self._on_edit)
        self._btnSave.clicked.connect(self._on_save)
        self._btnCancel.clicked.connect(self._on_cancel)
        self._btnDel.clicked.connect(self._on_del)
        self._tbl.itemSelectionChanged.connect(self._on_sel)
        self._set_ro(True); self._load()

    def _load(self):
        try:
            with db_manager.get_connection() as conn:
                rows = conn.execute(text(
                    "SELECT supplier_id, supplier_name, category, contact_name, is_active "
                    "FROM cims_supplier ORDER BY supplier_name"
                )).fetchall()
        except Exception as e:
            log.error("Supplier load: %s", e); return
        self._tbl.setRowCount(0)
        for row in rows:
            r = self._tbl.rowCount(); self._tbl.insertRow(r)
            for c, val in enumerate([
                str(row.supplier_id), row.supplier_name, row.category or "",
                row.contact_name or "", "Yes" if row.is_active else "No"
            ]):
                it = _cell(val)
                if c == 0: it.setData(Qt.UserRole, row.supplier_id)
                self._tbl.setItem(r, c, it)

    def _on_sel(self):
        sel = self._tbl.selectedItems()
        if not sel: return
        sid = self._tbl.item(sel[0].row(), 0).data(Qt.UserRole)
        self._current_id = sid
        try:
            with db_manager.get_connection() as conn:
                row = conn.execute(text(
                    "SELECT * FROM cims_supplier WHERE supplier_id=:id"
                ), {"id": sid}).fetchone()
            if not row: return
            self._txtName.setText(row.supplier_name or "")
            self._cmbCat.setCurrentIndex(self._cmbCat.findData(row.category))
            self._txtContact.setText(row.contact_name or "")
            self._txtPhone.setText(row.phone or "")
            self._txtEmail.setText(row.email or "")
            self._txtWeb.setText(row.website or "")
            self._chkActive.setChecked(bool(row.is_active))
            self._txtRem.setPlainText(row.remarks or "")
        except Exception as e:
            log.error("Supplier sel: %s", e)
        self._set_ro(True)

    def _set_ro(self, ro: bool):
        for w in [self._txtName, self._txtContact, self._txtPhone, self._txtEmail, self._txtWeb]:
            w.setReadOnly(ro)
        self._cmbCat.setEnabled(not ro); self._chkActive.setEnabled(not ro)
        self._txtRem.setReadOnly(ro)
        self._btnSave.setEnabled(not ro); self._btnCancel.setEnabled(not ro)
        has = self._current_id is not None
        self._btnEdit.setEnabled(ro and has); self._btnDel.setEnabled(ro and has)
        self._btnNew.setEnabled(ro)

    def _on_new(self):
        self._is_new = True; self._current_id = None
        for w in [self._txtName, self._txtContact, self._txtPhone, self._txtEmail, self._txtWeb]:
            w.clear()
        self._txtRem.clear(); self._chkActive.setChecked(True)
        self._set_ro(False); self._txtName.setFocus()

    def _on_edit(self): self._is_new = False; self._set_ro(False)

    def _on_cancel(self): self._is_new = False; self._set_ro(True)

    def _on_save(self):
        name = self._txtName.text().strip()
        if not name:
            show_message(self, "Validation", "Supplier name is required.")
            return
        p = {
            "name": name, "cat": self._cmbCat.currentData(),
            "con": self._txtContact.text().strip() or None,
            "ph":  self._txtPhone.text().strip() or None,
            "em":  self._txtEmail.text().strip() or None,
            "web": self._txtWeb.text().strip() or None,
            "act": 1 if self._chkActive.isChecked() else 0,
            "rem": self._txtRem.toPlainText().strip() or None,
            "now": datetime.now(), "user": getpass.getuser(),
        }
        try:
            with db_manager.get_connection() as conn:
                if self._is_new:
                    conn.execute(text("""
                        INSERT INTO cims_supplier
                          (supplier_name,category,contact_name,phone,email,website,
                           is_active,remarks,createdatestamp,createuserstamp,
                           modifdatestamp,modifuserstamp)
                        VALUES
                          (:name,:cat,:con,:ph,:em,:web,
                           :act,:rem,:now,:user,:now,:user)
                    """), p)
                else:
                    p["id"] = self._current_id
                    conn.execute(text("""
                        UPDATE cims_supplier SET
                          supplier_name=:name,category=:cat,contact_name=:con,
                          phone=:ph,email=:em,website=:web,is_active=:act,remarks=:rem,
                          modifdatestamp=:now,modifuserstamp=:user
                        WHERE supplier_id=:id
                    """), p)
                conn.commit()
            self._is_new = False; self._set_ro(True); self._load()
        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))

    def _on_del(self):
        if QMessageBox.question(
            self, "Delete", "Delete this supplier?", QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return
        try:
            with db_manager.get_connection() as conn:
                conn.execute(text(
                    "DELETE FROM cims_supplier WHERE supplier_id=:id"
                ), {"id": self._current_id}); conn.commit()
            self._current_id = None; self._load(); self._set_ro(True)
        except Exception as e:
            QMessageBox.critical(self, "Delete Error", str(e))


# ── Top-level widget ──────────────────────────────────────────────────────────

class CIMSWidget(QWidget):
    """Top-level CIMS widget embedded in the Settings sidebar stack."""

    def __init__(self, parent=None):
        super().__init__(parent)
        v = QVBoxLayout(self)
        v.setContentsMargins(6, 6, 6, 6)

        # toolbar
        tb = QHBoxLayout()
        lbl = QLabel("Consumables Inventory Management")
        lbl.setStyleSheet("font-size:14px;font-weight:bold;color:#37474F;")
        tb.addWidget(lbl)
        tb.addStretch(1)
        self._btnSuppliers = QPushButton("Suppliers…")
        self._btnSuppliers.setStyleSheet(_BTN_SS)
        self._btnSuppliers.clicked.connect(self._open_suppliers)
        tb.addWidget(self._btnSuppliers)
        tb.addWidget(make_help_button(self, "cims"))
        v.addLayout(tb)

        self._tabs = QTabWidget()
        self._inv_panel  = CIMSInventoryPanel()
        self._log_panel  = CIMSUsagePanel()
        self._rep_panel  = CIMSReportPanel()
        self._tabs.addTab(self._inv_panel,  "Inventory")
        self._tabs.addTab(self._log_panel,  "Usage Log")
        self._tabs.addTab(self._rep_panel,  "Reports")
        v.addWidget(self._tabs, 1)

    def _open_suppliers(self):
        dlg = _SupplierManagementDialog(self)
        dlg.exec_()
        # reload supplier list in inventory panel after editing
        self._inv_panel._load_suppliers()

    def get_alert_counts(self) -> dict:
        return self._inv_panel.get_alert_counts()
