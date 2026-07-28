"""
submission_report_gui.py — Submission Results Report for IsoWorks.
Generates a per-submission analytical results report from FinalValue,
exportable to PDF (via QPrinter) and Excel (via openpyxl).
Entry point: SubmissionReportDialog(submission_ids=[...], parent=...)

Report is pivoted: one row per sample, one column per parameter.
"""
from __future__ import annotations
import logging
from collections import defaultdict
from datetime import datetime
from typing import List

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTextBrowser, QFileDialog, QMessageBox, QProgressBar, QCheckBox,
    QGroupBox, QFormLayout, QLineEdit
)
from PyQt5.QtGui import QTextDocument
from PyQt5.QtCore import Qt
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

from db_core import db_manager
from sqlalchemy import text

# ---------------------------------------------------------------------------
_HEADER_COLOR = "#2D4A8A"
_ALT_ROW      = "#F3F7FA"

_STATUS_COLORS = {
    "Quantifiable":            "#C8E6C9",
    "Qualitative":             "#FFF9C4",
    "Below LC":                "#EEEEEE",
    "Below LC (Not Detected)": "#EEEEEE",
}

# LLD priority: lower = worse (for row-level colouring)
_LLD_PRIORITY = {"Below LC": 0, "Qualitative": 1, "Quantifiable": 2, "": 3}


def _fmt(val, nd=2):
    try:
        if val is None: return "—"
        return f"{float(val):.{nd}f}"
    except Exception as e:

        logging.warning(f"Exception caught: {e}"); return str(val)


def _lld_label(status):
    if status == 0:   return "Quantifiable"
    elif status == 1: return "Qualitative"
    else:             return "Below LC"


def _fetch_submission_rows(conn, sid):
    return conn.execute(text("""
        SELECT
            samp.Prefix || '-' || CAST(samp.SampleID AS TEXT)  AS SampleCode,
            samp.sName                                          AS SampleName,
            samp.CollectionDate,
            COALESCE(m.MeasurableName, m.ParameterLabel)        AS Parameter,
            fv.fValue,
            fv.fValueUnc,
            mu.ShortName                                        AS Unit,
            fv.LLDstatus,
            fv.RejectFlag,
            fv.AnalysisID,
            COALESCE(w.WorkflowName, '')                        AS JobName,
            samp.SampleID
        FROM FinalValue fv
        JOIN Analysis  a   ON fv.AnalysisID  = a.AnalysisID
        JOIN Sample   samp ON samp.SampleID  = a.SampleID
                           AND samp.Prefix   = a.Prefix
        JOIN Measurables m ON m.MeasurableID = fv.analyteid
        LEFT JOIN MeasurementUnit mu  ON mu.UnitID    = fv.MeasurableUnit
        LEFT JOIN Workflow         w   ON w.WorkflowID = a.WorkflowID
        WHERE samp.SubmissionID = :sid
          AND (fv.RejectFlag IS NULL OR fv.RejectFlag = false)
        ORDER BY samp.SampleID, COALESCE(m.MeasurableName, m.ParameterLabel)
    """), {'sid': sid}).fetchall()


def _pivot_rows(rows):
    """
    Returns:
        params       : ordered list of distinct ParameterLabel strings
        units        : dict {param: unit_string}
        sample_order : ordered list of (SampleCode, SampleName, CollectionDate)
        data         : dict {sample_key: {param: (fValue, fValueUnc, lld_label)}}
        analysis_ids : dict {sample_key: sorted list of AnalysisIDs}
        job_names    : dict {sample_key: sorted list of JobName strings}
    """
    params_ordered = []
    params_seen = set()
    units = {}
    sample_order = []
    sample_seen = set()
    data = defaultdict(dict)
    analysis_ids = defaultdict(set)
    job_names = defaultdict(set)

    for r in rows:
        param = r.Parameter or ''
        unit  = r.Unit or ''
        lld   = _lld_label(r.LLDstatus) if r.LLDstatus is not None else ''

        if param not in params_seen:
            params_seen.add(param)
            params_ordered.append(param)
            units[param] = unit

        key = (r.SampleCode, r.SampleName or '', str(r.CollectionDate or '')[:10])
        if key not in sample_seen:
            sample_seen.add(key)
            sample_order.append(key)

        data[key][param] = (_fmt(r.fValue, 3), _fmt(r.fValueUnc, 3), lld)
        if r.AnalysisID is not None:
            analysis_ids[key].add(r.AnalysisID)
        if r.JobName:
            job_names[key].add(r.JobName)

    # Convert sets to sorted lists
    analysis_ids = {k: sorted(v) for k, v in analysis_ids.items()}
    job_names    = {k: sorted(v) for k, v in job_names.items()}

    return params_ordered, units, sample_order, data, analysis_ids, job_names


def build_report_html(submission_ids: List[int]) -> str:
    """Query DB and return a full HTML report (wide/pivoted format)."""
    sections = []

    with db_manager.get_connection() as conn:
        for sid in submission_ids:
            # --- Submission header ---
            sub = conn.execute(text("""
                SELECT s.SubmissionID, s.SubmissionName, s.SubmissionDate,
                       s.SubmissionSite,
                       md.medianame                                AS MediaName,
                       w.WorkflowName,
                       COALESCE(cust.LastName || ', ' || cust.FirstName, '')        AS Client,
                       COALESCE(tech.LastName || ', ' || tech.FirstMiddleName, '') AS Officer
                FROM Submission s
                LEFT JOIN Media    md   ON md.MediaID   = s.MediaID
                LEFT JOIN Workflow w    ON w.WorkflowID = s.RequestedWorkflow
                LEFT JOIN Customer cust ON cust.CustomerID = s.CustomerID
                LEFT JOIN Employee tech ON tech.EmployeeID = s.TechnicalOfficer
                WHERE s.SubmissionID = :sid
            """), {'sid': sid}).fetchone()

            if not sub:
                continue

            sub_date = str(sub.SubmissionDate or '—')[:10]
            html = f"""
            <h2 style='color:{_HEADER_COLOR};border-bottom:2px solid {_HEADER_COLOR};
                       padding-bottom:4px;margin-top:24px;'>
                Submission {sub.SubmissionID} — {sub.SubmissionName or ''}
            </h2>
            <table cellpadding='5' cellspacing='0' style='font-size:13px;margin-bottom:12px;'>
              <tr><td><b>Date:</b></td><td>{sub_date}</td>
                  <td width='30'></td>
                  <td><b>Client:</b></td><td>{sub.Client or '—'}</td></tr>
              <tr><td><b>Media:</b></td><td>{sub.MediaName or '—'}</td>
                  <td></td>
                  <td><b>Workflow:</b></td><td>{sub.WorkflowName or '—'}</td></tr>
              <tr><td><b>Site:</b></td><td colspan='4'>{sub.SubmissionSite or '—'}</td></tr>
            </table>
            """

            rows = _fetch_submission_rows(conn, sid)

            if not rows:
                html += "<p><i>No results available for this submission.</i></p>"
                sections.append(html)
                continue

            params, units, sample_order, data, analysis_ids, job_names = _pivot_rows(rows)

            # --- Build dynamic header ---
            fixed_cols = ["Sample", "Name", "Collection", "Analysis ID(s)", "Job(s)"]
            html += f"""
            <table border='0' cellspacing='0' cellpadding='7'
                   style='border-collapse:collapse;width:100%;font-size:13px;'>
              <thead>
                <tr style='background:{_HEADER_COLOR};color:white;'>
            """
            for col in fixed_cols:
                html += f"<th align='left'>{col}</th>"
            for p in params:
                u = units.get(p, '')
                label = f"{p}<br><span style='font-weight:normal;font-size:11px;'>({u})</span>" if u else p
                html += f"<th align='right'>{label}</th>"
            html += "</tr>\n"

            # Sub-header row
            html += f"<tr style='background:{_HEADER_COLOR};color:#cce0ff;font-size:11px;'>"
            for _ in fixed_cols:
                html += "<th></th>"
            for p in params:
                html += "<th align='right'><i>val ± unc</i></th>"
            html += "</tr>\n</thead><tbody>\n"

            # --- Data rows ---
            for i, key in enumerate(sample_order):
                sample_code, sample_name, col_date = key
                sample_data = data[key]
                aids  = ', '.join(str(a) for a in analysis_ids.get(key, []))
                jobs  = ', '.join(job_names.get(key, []))

                # Worst-case LLD for row colour
                worst_lld = ''
                for p in params:
                    cell = sample_data.get(p)
                    if cell:
                        lld = cell[2]
                        if _LLD_PRIORITY.get(lld, 3) < _LLD_PRIORITY.get(worst_lld, 3):
                            worst_lld = lld
                bg = _STATUS_COLORS.get(worst_lld, _ALT_ROW if i % 2 else 'white')

                html += f"<tr style='background:{bg};'>"
                html += f"<td>{sample_code}</td>"
                html += f"<td>{sample_name}</td>"
                html += f"<td>{col_date}</td>"
                html += f"<td style='font-size:11px;color:#444;'>{aids}</td>"
                html += f"<td style='font-size:11px;'>{jobs}</td>"

                for p in params:
                    cell = sample_data.get(p)
                    if cell:
                        val, unc, lld = cell
                        lld_sup = f" <sup style='color:#888;font-size:10px;'>{lld[0]}</sup>" if lld else ''
                        html += f"<td align='right'>{val} ± {unc}{lld_sup}</td>"
                    else:
                        html += "<td align='right' style='color:#bbb;'>—</td>"
                html += "</tr>\n"

            html += "</tbody></table>"

            # Legend
            html += f"""
            <p style='font-size:11px;color:#555;margin-top:6px;'>
              LLD status superscript: <b>Q</b>=Quantifiable &nbsp;
              <b>q</b>=Qualitative &nbsp; <b>B</b>=Below LC
            </p>
            """
            sections.append(html)

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    header = f"""
    <html><head><meta charset='UTF-8'>
    <style>
      body {{ font-family: Arial, sans-serif; font-size: 13px; margin: 24px; }}
      h1   {{ color: {_HEADER_COLOR}; font-size: 20px; margin-bottom: 4px; }}
      h2   {{ font-size: 15px; }}
      table {{ font-size: 13px; }}
      th, td {{ padding: 6px 10px; }}
      thead tr:first-child th {{ border-bottom: 1px solid {_HEADER_COLOR}; }}
    </style></head><body>
    <h1>Analytical Results Report</h1>
    <p style='color:#666;font-size:12px;'>Generated: {generated} &nbsp;|&nbsp;
       Submissions: {', '.join(str(s) for s in submission_ids)}</p>
    <hr style='border:1px solid {_HEADER_COLOR};'>
    """
    footer = "</body></html>"
    return header + "".join(sections) + footer


class SubmissionReportDialog(QDialog):
    """
    Preview + export report for one or more submissions.
    submission_ids: list of SubmissionID integers.
    """
    def __init__(self, submission_ids: List[int], parent=None):
        super().__init__(parent)
        self.submission_ids = submission_ids
        self.setWindowTitle("Submission Results Report")
        self.resize(1100, 800)
        self.setAttribute(Qt.WA_DeleteOnClose)

        layout = QVBoxLayout(self)

        # Toolbar
        toolbar = QHBoxLayout()
        self.btnPrint   = QPushButton("Print / Save PDF…")
        self.btnExcel   = QPushButton("Export to Excel…")
        self.btnClose   = QPushButton("Close")
        self.btnRefresh = QPushButton("Refresh")
        self.lblInfo    = QLabel()
        toolbar.addWidget(self.btnPrint)
        toolbar.addWidget(self.btnExcel)
        toolbar.addWidget(self.btnRefresh)
        toolbar.addStretch()
        toolbar.addWidget(self.lblInfo)
        toolbar.addWidget(self.btnClose)
        layout.addLayout(toolbar)

        # Progress
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setMaximumHeight(6)
        layout.addWidget(self.progress)

        # Preview pane
        self.browser = QTextBrowser()
        self.browser.setOpenExternalLinks(False)
        layout.addWidget(self.browser, 1)

        # Signals
        self.btnClose.clicked.connect(self.close)
        self.btnRefresh.clicked.connect(self._load)
        self.btnPrint.clicked.connect(self._print)
        self.btnExcel.clicked.connect(self._export_excel)

        self._load()

    def _load(self):
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        try:
            self._html = build_report_html(self.submission_ids)
            self.browser.setHtml(self._html)
            self.lblInfo.setText(f"{len(self.submission_ids)} submission(s) loaded")
        except Exception as e:
            logging.error(f"Report load failed: {e}", exc_info=True)
            self.browser.setHtml(f"<b>Error loading report:</b> {e}")
        finally:
            self.progress.setVisible(False)

    def _print(self):
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.A4)
        dlg = QPrintDialog(printer, self)
        if dlg.exec_() != QPrintDialog.Accepted:
            return
        doc = QTextDocument()
        doc.setHtml(self._html)
        doc.print_(printer)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Missing library",
                                 "openpyxl is required for Excel export.\n"
                                 "Install with: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", "results_report.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            hdr_font  = Font(bold=True, color="FFFFFF", size=12)
            hdr_fill  = PatternFill("solid", fgColor="2D4A8A")
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin      = Side(style="thin", color="CCCCCC")
            bdr       = Border(bottom=thin)

            fill_alt   = PatternFill("solid", fgColor="F3F7FA")
            fill_green = PatternFill("solid", fgColor="C8E6C9")
            fill_yel   = PatternFill("solid", fgColor="FFF9C4")
            fill_grey  = PatternFill("solid", fgColor="EEEEEE")

            with db_manager.get_connection() as conn:
                for sid_idx, sid in enumerate(self.submission_ids):
                    if sid_idx > 0:
                        ws.append([])  # blank separator row between submissions

                    rows = _fetch_submission_rows(conn, sid)
                    if not rows:
                        ws.append([f"Submission {sid} — no results"])
                        continue

                    params, units, sample_order, data, analysis_ids, job_names = _pivot_rows(rows)

                    # Header row
                    fixed_hdrs = ["Submission", "Sample", "Sample Name", "Collection Date", "Analysis ID(s)", "Job(s)"]
                    param_hdrs = [f"{p} ({units.get(p,'')})" if units.get(p) else p for p in params]
                    all_hdrs = fixed_hdrs + param_hdrs
                    ws.append(all_hdrs)
                    hdr_row_idx = ws.max_row
                    for cell in ws[hdr_row_idx]:
                        cell.font      = hdr_font
                        cell.fill      = hdr_fill
                        cell.alignment = hdr_align

                    # Unit sub-header
                    unit_row = [''] * len(fixed_hdrs) + [f"val ± unc  [LLD]" for _ in params]
                    ws.append(unit_row)
                    sub_hdr_row_idx = ws.max_row
                    sub_font = Font(italic=True, color="FFFFFF", size=10)
                    sub_fill = PatternFill("solid", fgColor="3A5CA0")
                    for cell in ws[sub_hdr_row_idx]:
                        cell.font = sub_font
                        cell.fill = sub_fill
                        cell.alignment = Alignment(horizontal="center")

                    # Data rows
                    for i, key in enumerate(sample_order):
                        sample_code, sample_name, col_date = key
                        sample_data = data[key]
                        aids = ', '.join(str(a) for a in analysis_ids.get(key, []))
                        jobs = ', '.join(job_names.get(key, []))

                        # Worst-case LLD for row fill
                        worst_lld = ''
                        for p in params:
                            cell_data = sample_data.get(p)
                            if cell_data:
                                lld = cell_data[2]
                                if _LLD_PRIORITY.get(lld, 3) < _LLD_PRIORITY.get(worst_lld, 3):
                                    worst_lld = lld

                        row_vals = [sid, sample_code, sample_name, col_date, aids, jobs]
                        for p in params:
                            cell_data = sample_data.get(p)
                            if cell_data:
                                val, unc, lld = cell_data
                                row_vals.append(f"{val} ± {unc}  [{lld}]" if lld else f"{val} ± {unc}")
                            else:
                                row_vals.append('')
                        ws.append(row_vals)

                        if worst_lld == "Quantifiable":    fill = fill_green
                        elif worst_lld == "Qualitative":   fill = fill_yel
                        elif "Below" in worst_lld:         fill = fill_grey
                        else:                              fill = fill_alt if i % 2 else None
                        if fill:
                            for cell in ws[ws.max_row]:
                                cell.fill = fill

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Saved to:\n{path}")

        except Exception as e:
            logging.error(f"Excel export failed: {e}", exc_info=True)
            QMessageBox.critical(self, "Export failed", str(e))
