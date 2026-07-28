"""
ngam_ng_parser.py
=================
Parser for Noble Gas mass spectrometer sequence XLSM workbooks.

Entry point: parse_ng_sequence(folder_path) -> NGSequenceParsed

folder_path should be the top-level sequence folder containing:
  - one *.xlsm main workbook file
  - SequenceFitResults/ subdirectory with per-isotope *.xlsm files
"""
from __future__ import annotations

import glob
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional

log = logging.getLogger(__name__)

# ── name parsing regexes ───────────────────────────────────────────────────────
_SUBSAMPLE_RE = re.compile(
    r'\s*SubSamplePort(\d+)_SID(\d+)_SSID(.+)', re.IGNORECASE
)


# ─────────────────────────────── dataclasses ──────────────────────────────────

@dataclass
class NGInletEntry:
    """One row from SeqFitPipStg."""
    inlet_num: int
    name: str                    # raw name from sheet
    position_type: str           # "blank"|"standard"|"air_ref"|"sample"|"invalid"
    port_num: Optional[int]      # parsed from SubSamplePortX
    sample_id: Optional[int]     # parsed from SIDYYY
    subsample_id: Optional[str]  # parsed from SSIDZZZ
    size_hint: Optional[float]


@dataclass
class NGIsotopeSignals:
    """Per-inlet signals for one isotope from SequenceFitResults."""
    isotope: str              # e.g. "He", "4He", "20Ne", etc.
    inlet_num: int
    inlet_time: Optional[float] = None
    blank_fit: Optional[float] = None
    blank_corrected: Optional[float] = None   # col 33 (0-indexed: 32)
    repro_corrected: Optional[float] = None   # col 36 (0-indexed: 35)
    lin_corrected: Optional[float] = None     # col 37 (0-indexed: 36)
    blank_fit_upper: Optional[float] = None   # BlankFit 95% CI upper (col 3, 0-based)
    blank_fit_lower: Optional[float] = None   # BlankFit 95% CI lower (col 4, 0-based)
    blank_signal: Optional[float] = None      # raw blank measurement signal (col 5)
    repro_fit: Optional[float] = None         # drift correction factor (col 10)
    repro_fit_upper: Optional[float] = None   # ReproFit 95% CI upper (col 11)
    repro_fit_lower: Optional[float] = None   # ReproFit 95% CI lower (col 12)
    repro_signal: Optional[float] = None      # repro measurement signal (col 13)
    lin_fit_x: Optional[float] = None         # linearity fit X coordinate (col 18)
    lin_fit: Optional[float] = None           # linearity correction factor (col 19)
    lin_fit_upper: Optional[float] = None     # LinFit 95% CI upper (col 20)
    lin_fit_lower: Optional[float] = None     # LinFit 95% CI lower (col 21)
    lin_signal: Optional[float] = None        # lin measurement signal (col 22)
    cal_x: Optional[float] = None            # known calibration concentration ccSTP (col 27)
    cal_y: Optional[float] = None            # calibration standard signal (col 28)


@dataclass
class NGFinalResult:
    """One row from Final Results 'Mass Spectrometers' section."""
    srg: str                       # Sample Reference Group name
    he_ccstp_g: Optional[float]
    he_err: Optional[float]
    he_err_pct: Optional[float]
    ne_ccstp_g: Optional[float]
    ne_err: Optional[float]
    ne_err_pct: Optional[float]
    ar_ccstp_g: Optional[float]
    ar_err: Optional[float]
    ar_err_pct: Optional[float]
    kr_ccstp_g: Optional[float]
    kr_err: Optional[float]
    kr_err_pct: Optional[float]
    xe_ccstp_g: Optional[float]
    xe_err: Optional[float]
    xe_err_pct: Optional[float]
    ngt: Optional[float]           # Noble Gas Temperature (from first section)
    excess_air: Optional[float]
    chi_square: Optional[float]


@dataclass
class NGSequenceParsed:
    folder_path: str
    main_file: str                                      # filename of main .xlsm
    run_id_hint: Optional[str]                          # from filename if parseable
    pipeline: List[NGInletEntry]                        # from SeqFitPipStg
    isotope_signals: Dict[str, List[NGIsotopeSignals]]  # isotope → list per inlet
    final_results: List[NGFinalResult]                  # from Final Results sheet
    n_inlets: int
    n_samples: int     # count of type==sample
    n_blanks: int
    n_standards: int


# ─────────────────────────────── helpers ──────────────────────────────────────

def _safe_float(val) -> Optional[float]:
    """Convert cell value to float; return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _classify_name(name: str) -> tuple:
    """
    Classify a SeqFitPipStg name string.

    Returns (position_type, port_num, sample_id, subsample_id)
    where position_type is one of: blank, standard, air_ref, sample, invalid
    """
    if name is None:
        return "invalid", None, None, None
    name_stripped = str(name).strip()
    if not name_stripped or name_stripped == "None":
        return "invalid", None, None, None

    # SubSamplePort
    m = _SUBSAMPLE_RE.match(name_stripped)
    if m:
        port_num = int(m.group(1))
        sample_id = int(m.group(2))
        subsample_id = m.group(3).strip()
        return "sample", port_num, sample_id, subsample_id

    nl = name_stripped.lower()
    # Blank check
    if nl == "blank" or "blank" in nl:
        return "blank", None, None, None
    # Standard / Spike → repro reference
    if "standard" in nl or "spike" in nl:
        # Air standards
        if "air" in nl:
            return "air_ref", None, None, None
        return "standard", None, None, None
    # Air reference
    if "air" in nl:
        return "air_ref", None, None, None

    return "sample", None, None, None


# ─────────────────────────────── SeqFitPipStg parser ──────────────────────────

def _parse_seqfitpipstg(ws) -> List[NGInletEntry]:
    """Parse the SeqFitPipStg sheet. Rows 1-4 are header, row 5 column headers,
    rows 6+ are data (1-indexed per openpyxl read_only)."""
    entries: List[NGInletEntry] = []

    data_started = False
    for row in ws.iter_rows(values_only=True):
        col_a = row[0] if len(row) > 0 else None

        # Skip until we find the data rows (col A is an integer)
        if not data_started:
            # Check if col A is "InletNumber" (header row) or an integer
            if col_a is None:
                continue
            if str(col_a).strip() == "InletNumber":
                data_started = True
                continue
            # Try as integer
            try:
                int(col_a)
                data_started = True
                # fall through to process this row
            except (TypeError, ValueError):
                continue

        # data row
        if col_a is None:
            continue
        try:
            inlet_num = int(col_a)
        except (TypeError, ValueError):
            continue

        name = row[1] if len(row) > 1 else None
        name_str = str(name).strip() if name is not None else "None"

        # Stop at template placeholder rows (None name, high inlet number)
        if name_str == "None" and inlet_num > 30:
            break

        size_hint = _safe_float(row[2]) if len(row) > 2 else None

        ptype, port_num, sample_id, subsample_id = _classify_name(name_str)

        entries.append(NGInletEntry(
            inlet_num=inlet_num,
            name=name_str,
            position_type=ptype,
            port_num=port_num,
            sample_id=sample_id,
            subsample_id=subsample_id,
            size_hint=size_hint,
        ))

    return entries


# ─────────────────────────────── Final Results parser ─────────────────────────

def _parse_final_results(ws) -> List[NGFinalResult]:
    """
    Parse both sections of the Final Results sheet:
      - First section (row 2 header): SRG + He/Ne/Ar + NGT/ExcessAir/ChiSq
      - 'Mass Spectrometers' section: SRG + He/Ne/Ar/Kr/Xe

    Merges both into NGFinalResult objects keyed by SRG.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # Build index of rows where col 0 contains useful info
    # ── Find first section header (row index with 'SRG' in col 0) ─────────────
    first_section_header_idx = None
    mass_spec_section_idx = None

    for i, row in enumerate(all_rows):
        if not row:
            continue
        cell0 = str(row[0]).strip() if row[0] is not None else ""
        # Look for 'Mass Spectrometers' marker
        if "mass spectrometer" in cell0.lower():
            mass_spec_section_idx = i
        # Look for first section header (SRG header row)
        if cell0 == "SRG" and first_section_header_idx is None:
            first_section_header_idx = i

    # ── Parse first section for NGT/ExcessAir/ChiSq ───────────────────────────
    # col indices (0-based): 0=SRG, 15=NGT, 16=ExcessAir, 17=ExAirInCu, 18=ChiSq
    first_section_data: Dict[str, tuple] = {}
    if first_section_header_idx is not None:
        for row in all_rows[first_section_header_idx + 1:]:
            if not row:
                continue
            # Stop at Mass Spectrometers section or empty SRG
            srg = row[0] if len(row) > 0 else None
            if srg is None:
                continue
            srg_str = str(srg).strip()
            if not srg_str or srg_str == "Mass Spectrometers":
                if mass_spec_section_idx is not None:
                    break
                continue
            ngt       = _safe_float(row[15]) if len(row) > 15 else None
            excess_air = _safe_float(row[16]) if len(row) > 16 else None
            chi_sq    = _safe_float(row[18]) if len(row) > 18 else None
            first_section_data[srg_str] = (ngt, excess_air, chi_sq)

    # ── Parse Mass Spectrometers section ──────────────────────────────────────
    results: List[NGFinalResult] = []
    if mass_spec_section_idx is None:
        log.warning("Final Results: 'Mass Spectrometers' section not found")
        return results

    # The row after mass_spec_section_idx is the column headers row
    # Data rows start after that
    data_start = mass_spec_section_idx + 2  # skip header+section label row
    for row in all_rows[data_start:]:
        if not row:
            continue
        srg = row[0] if len(row) > 0 else None
        if srg is None:
            break
        srg_str = str(srg).strip()
        if not srg_str:
            break

        def _f(idx):
            return _safe_float(row[idx]) if len(row) > idx else None

        ngt, excess_air, chi_sq = first_section_data.get(srg_str, (None, None, None))

        results.append(NGFinalResult(
            srg=srg_str,
            he_ccstp_g=_f(1),
            he_err=_f(2),
            he_err_pct=_f(3),
            ne_ccstp_g=_f(4),
            ne_err=_f(5),
            ne_err_pct=_f(6),
            ar_ccstp_g=_f(7),
            ar_err=_f(8),
            ar_err_pct=_f(9),
            kr_ccstp_g=_f(10),
            kr_err=_f(11),
            kr_err_pct=_f(12),
            xe_ccstp_g=_f(13),
            xe_err=_f(14),
            xe_err_pct=_f(15),
            ngt=ngt,
            excess_air=excess_air,
            chi_square=chi_sq,
        ))

    return results


# ─────────────────────────────── SequenceFitResults parser ────────────────────

def _parse_isotope_file(filepath: str, isotope_name: str) -> List[NGIsotopeSignals]:
    """
    Parse a single SequenceFitResults/*.xlsm file.

    Col mapping (0-indexed):
      0  = InletNumber
      1  = InletTime
      2  = BlankFit
      32 = BlankCorrected
      35 = ReproCorrected
      36 = LinearityCorrected
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            filepath, read_only=True, data_only=True, keep_vba=False
        )
    except Exception as exc:
        log.warning("Cannot open isotope file %s: %s", filepath, exc)
        return []

    signals: List[NGIsotopeSignals] = []
    try:
        ws = wb.active
        data_started = False
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            col0 = row[0]
            if not data_started:
                # Look for the InletNumber header row (row 5, 1-indexed)
                if col0 is not None and str(col0).strip() == "InletNumber":
                    data_started = True
                continue

            # Data row
            if col0 is None:
                continue
            try:
                inlet_num = int(col0)
            except (TypeError, ValueError):
                continue

            def _g(idx):
                return _safe_float(row[idx]) if len(row) > idx else None

            signals.append(NGIsotopeSignals(
                isotope=isotope_name,
                inlet_num=inlet_num,
                inlet_time=_g(1),
                blank_fit=_g(2),
                blank_corrected=_g(32),
                repro_corrected=_g(35),
                lin_corrected=_g(36),
                blank_fit_upper=_g(3),
                blank_fit_lower=_g(4),
                blank_signal=_g(5),
                repro_fit=_g(10),
                repro_fit_upper=_g(11),
                repro_fit_lower=_g(12),
                repro_signal=_g(13),
                lin_fit_x=_g(18),
                lin_fit=_g(19),
                lin_fit_upper=_g(20),
                lin_fit_lower=_g(21),
                lin_signal=_g(22),
                cal_x=_g(27),
                cal_y=_g(28),
            ))
    except Exception as exc:
        log.warning("Error reading isotope file %s: %s", filepath, exc)
    finally:
        try:
            wb.close()
        except Exception as e:

            logging.warning(f"Exception caught: {e}")

    return signals


# ─────────────────────────────── main entry point ─────────────────────────────

def parse_ng_sequence(folder_path: str) -> NGSequenceParsed:
    """
    Parse a Noble Gas sequence folder.

    Parameters
    ----------
    folder_path : str
        Top-level sequence folder path.

    Returns
    -------
    NGSequenceParsed
    """
    import openpyxl

    folder_path = os.path.normpath(folder_path)

    # ── 1. Find main workbook ──────────────────────────────────────────────────
    xlsm_files = glob.glob(os.path.join(folder_path, "*.xlsm"))
    if not xlsm_files:
        raise FileNotFoundError(f"No .xlsm file found in: {folder_path}")
    main_file = xlsm_files[0]
    main_filename = os.path.basename(main_file)

    # Try to extract run ID hint from filename (e.g. "10565-2024_08_14-09_40")
    run_id_hint: Optional[str] = None
    basename = os.path.splitext(main_filename)[0]
    m = re.match(r'^(\d+)-', basename)
    if m:
        run_id_hint = basename

    log.info("Opening main workbook: %s", main_file)

    # ── 2. Parse main workbook ────────────────────────────────────────────────
    pipeline: List[NGInletEntry] = []
    final_results: List[NGFinalResult] = []

    try:
        wb_main = openpyxl.load_workbook(
            main_file, read_only=True, data_only=True, keep_vba=False
        )
        sheet_names_lower = {s.lower(): s for s in wb_main.sheetnames}

        # SeqFitPipStg
        sfps_key = next(
            (k for k in sheet_names_lower if "seqfitpipstg" in k), None
        )
        if sfps_key:
            ws_sfps = wb_main[sheet_names_lower[sfps_key]]
            try:
                pipeline = _parse_seqfitpipstg(ws_sfps)
            except Exception as exc:
                log.warning("Error parsing SeqFitPipStg: %s", exc)
        else:
            log.warning("Sheet 'SeqFitPipStg' not found in %s", main_filename)

        # Final Results
        fr_key = next(
            (k for k in sheet_names_lower if "final result" in k), None
        )
        if fr_key:
            ws_fr = wb_main[sheet_names_lower[fr_key]]
            try:
                final_results = _parse_final_results(ws_fr)
            except Exception as exc:
                log.warning("Error parsing Final Results: %s", exc)
        else:
            log.warning("Sheet 'Final Results' not found in %s", main_filename)

        wb_main.close()
    except Exception as exc:
        log.error("Failed to open main workbook %s: %s", main_file, exc)
        raise

    # ── 3. Parse SequenceFitResults/ ──────────────────────────────────────────
    sfr_dir = os.path.join(folder_path, "SequenceFitResults")
    isotope_signals: Dict[str, List[NGIsotopeSignals]] = {}

    if os.path.isdir(sfr_dir):
        iso_files = glob.glob(os.path.join(sfr_dir, "*.xlsm"))
        for iso_path in sorted(iso_files):
            iso_name = os.path.splitext(os.path.basename(iso_path))[0]
            sigs = _parse_isotope_file(iso_path, iso_name)
            if sigs:
                isotope_signals[iso_name] = sigs
                log.info("Parsed %d rows from %s", len(sigs), iso_name)
    else:
        log.warning("SequenceFitResults directory not found: %s", sfr_dir)

    # ── 4. Compute summary counts ─────────────────────────────────────────────
    n_inlets   = len(pipeline)
    n_samples  = sum(1 for e in pipeline if e.position_type == "sample")
    n_blanks   = sum(1 for e in pipeline if e.position_type == "blank")
    n_standards = sum(1 for e in pipeline if e.position_type == "standard")

    return NGSequenceParsed(
        folder_path=folder_path,
        main_file=main_filename,
        run_id_hint=run_id_hint,
        pipeline=pipeline,
        isotope_signals=isotope_signals,
        final_results=final_results,
        n_inlets=n_inlets,
        n_samples=n_samples,
        n_blanks=n_blanks,
        n_standards=n_standards,
    )
